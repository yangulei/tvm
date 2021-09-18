import openpyxl
import codecs
from openpyxl.utils import get_column_letter
import os
 
def txt_to_xlsx(filename, sheetname, outfile):
    
    first = os.path.exists(outfile)
    fr = codecs.open(filename,'r')
    if not first:
        wb = openpyxl.Workbook()
    else:
        wb=openpyxl.load_workbook(outfile)
    ws = wb.active
    ws = wb.create_sheet()
    ws.title = sheetname
    row = 0
    cnt = 0
    for line in fr:
        line = line.strip()
        line = line.split(',')
        row +=1
        col = 0
        for j in range(len(line)):
            col +=1
            print (line[j])
            ws.cell(column = col,row = row,value = line[j].format(get_column_letter(col)))
        
    wb.save(outfile)


def read_xlsx(filename):
    #载入文件
    wb = openpyxl.load_workbook(filename)
    #获取Sheet1工作表
    ws = wb.get_sheet_by_name('Sheet1')
    #按行读取
    for row in ws.rows:
        for cell in row:
            print (cell.value)
    #按列读
    for col in ws.columns:
        for cell in col:
            print (cell.value)
 
if __name__=='__main__':
    #rootdir = "/home2/zhangya9/tvm/tutorials/experiment_res/0906"
    outfileExcel = '/home2/zhangya9/tvm/tutorials/experiment_res/0917/opt_byoc_v1.7.xlsx'
    inputfileTxt = '/home2/zhangya9/tvm/tutorials/experiment_res/0917/opt_byoc_bs128_v1.7.txt'
    sheetname = inputfileTxt.split('/')[-1].split('.')[0]
    txt_to_xlsx(inputfileTxt, sheetname, outfileExcel)
    """
    for (dirpath,dirnames,filenames) in os.walk(rootdir):
        for filename in filenames:
            if os.path.splitext(filename)[1]=='.txt':
                inputfileTxt = os.path.join(rootdir, filename)
                sheetname = inputfileTxt.split('/')[-1].split('.')[0]
                txt_to_xlsx(inputfileTxt, sheetname, outfileExcel, first)
                first = False
    """
    
    # read_xlsx(outfileExcel)
